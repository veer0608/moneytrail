"""Statements in, one ledger out -- with a certificate that says it is complete.

Every other tool that turns a bank PDF into a spreadsheet can be wrong without
saying so. A wrapped narration swallows a row, a page break eats the last line
of a table, and what lands in the spreadsheet still looks like a spreadsheet.
The person importing it has no way to tell.

This module exists to make that failure impossible to receive silently. The
export carries the reconciliation with it: a :class:`Certificate` per source
file, naming the arithmetic that was checked, the SHA-256 of the exact bytes it
was checked against, and whether it held. An export that failed to reconcile is
still written -- refusing to hand over the data helps nobody -- but it is
written stamped ``NOT RECONCILED``, and the CLI exits non-zero.

The certificate is bound to a digest rather than to a filename on purpose. A
filename says which file was *meant*; a digest says which bytes were actually
read. Only the second survives being emailed around.
"""

from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterator, Sequence

from .models import CardStatement, Direction, Statement
from .money import Paise, format_paise
from .query import build_ledger
from .reconcile import is_tautological, reconcile, reconcile_card

#: The export layout. Date, narration and the debit/credit pair come first
#: because that is the shape every accounting import expects; provenance goes
#: at the end where it does not get in the way but is still there to audit.
COLUMNS: tuple[str, ...] = (
    "date",
    "value_date",
    "narration",
    "merchant",
    "category",
    "counterparty_type",
    "debit",
    "credit",
    "balance",
    "institution",
    "account",
    "source_file",
    "page",
    "row",
    "reconciled",
)

DIGEST_UNAVAILABLE = "unavailable"


def rupees(paise: Paise) -> Decimal:
    """Paise as an exact rupee ``Decimal``.

    Not a float. ``money.py`` bans them because the promise is "to the paisa",
    and a spreadsheet full of binary-rounded rupees cannot keep that promise
    the moment someone sums a column. ``Decimal`` is exact and both ``csv`` and
    ``openpyxl`` write it as a number rather than as text.
    """
    return Decimal(paise).scaleb(-2)


def digest_of(path: Path) -> str:
    """SHA-256 of the source file, or ``unavailable`` if it cannot be read.

    Synthetic statements built in memory have no bytes behind them, and that is
    a legitimate state rather than an error -- so this never raises. A missing
    digest weakens the certificate and says so in the text.
    """
    try:
        return hashlib.sha256(Path(path).read_bytes()).hexdigest()
    except OSError:
        return DIGEST_UNAVAILABLE


@dataclass(frozen=True)
class Certificate:
    """What was checked on one statement, and whether it held.

    This is the part of the export that the other converters cannot produce,
    so it is a first-class object rather than a formatted string: the web
    front-end renders it, the CLI prints it, and both read the same fields.
    """

    source: Path
    kind: str  # "bank" | "card"
    institution: str
    account_hint: str
    digest: str
    transactions: int
    reconciled: bool
    checks: tuple[str, ...]
    caveats: tuple[str, ...] = ()
    failures: tuple[str, ...] = ()
    period_start: date | None = None
    period_end: date | None = None
    opening: Paise | None = None
    credits: Paise | None = None
    debits: Paise | None = None
    computed_closing: Paise | None = None
    stated_closing: Paise | None = None

    @property
    def verdict(self) -> str:
        return "RECONCILED" if self.reconciled else "NOT RECONCILED"

    def render(self) -> str:
        """The certificate as plain text, for a sidecar file or a terminal."""
        period = f"{self.period_start or '?'} to {self.period_end or '?'}"
        lines = [
            f"  source          {Path(self.source).name}",
            f"  sha-256         {self.digest}",
            f"  {'card' if self.kind == 'card' else 'bank':<15} {self.institution or '-'}",
            f"  account         {self.account_hint or '-'}",
            f"  period          {period}",
            f"  transactions    {self.transactions}",
        ]
        if self.kind == "bank":
            lines += [
                f"  opening       {format_paise(self.opening or 0):>16}",
                f"  credits     + {format_paise(self.credits or 0):>16}",
                f"  debits      - {format_paise(self.debits or 0):>16}",
                f"  computed      {format_paise(self.computed_closing or 0):>16}",
                f"  statement     {format_paise(self.stated_closing or 0):>16}",
            ]
        else:
            lines += [
                f"  charged       {format_paise(self.debits or 0):>16}",
                f"  paid off      {format_paise(self.credits or 0):>16}",
            ]

        checks = ", ".join(self.checks) if self.checks else "none available"
        lines.append(f"  checks run      {checks}")
        lines.append(f"  VERDICT         {self.verdict}")
        lines += [f"    caveat: {c}" for c in self.caveats]
        lines += [f"    failure: {f}" for f in self.failures]
        return "\n".join(lines)


def certify(statement: Statement | CardStatement) -> Certificate:
    """Reconcile one statement and record the outcome as a certificate."""
    if isinstance(statement, CardStatement):
        return _certify_card(statement)
    return _certify_bank(statement)


def _certify_bank(statement: Statement) -> Certificate:
    result = reconcile(statement)

    caveats: list[str] = []
    checks = ["totals"]
    if result.chain_checked:
        checks.insert(0, "chain")
    else:
        caveats.append(
            "no running-balance column, so a fault could not be localised to a row"
        )
    if is_tautological(statement):
        caveats.append(
            "both endpoints were derived from the rows, so the first and last "
            "row are not independently checked"
        )

    return Certificate(
        source=statement.source,
        kind="bank",
        institution=statement.bank,
        account_hint=statement.account_hint,
        digest=digest_of(statement.source),
        transactions=len(statement.transactions),
        reconciled=result.ok,
        checks=tuple(checks),
        caveats=tuple(caveats),
        failures=tuple(d.describe() for d in result.discrepancies),
        period_start=statement.period_start,
        period_end=statement.period_end,
        opening=statement.opening_balance,
        credits=result.credits,
        debits=result.debits,
        computed_closing=result.computed_closing,
        stated_closing=statement.closing_balance,
    )


def _certify_card(statement: CardStatement) -> Certificate:
    result = reconcile_card(statement)

    caveats: list[str] = []
    if not result.checks:
        caveats.append(
            "the statement prints no totals, so the rows could not be checked "
            "against anything"
        )
    elif not result.verified:
        caveats.append(
            "only the summary box was checked; the transaction rows were not "
            "cross-checked against a stated total"
        )

    # No checks at all is not a pass. An unverifiable parse must never be
    # stamped RECONCILED, or the certificate means nothing on the statements
    # that need it most.
    reconciled = result.ok and bool(result.checks)

    return Certificate(
        source=statement.source,
        kind="card",
        institution=statement.issuer,
        account_hint=statement.account_hint,
        digest=digest_of(statement.source),
        transactions=len(statement.transactions),
        reconciled=reconciled,
        checks=result.checks,
        caveats=tuple(caveats),
        failures=tuple(d.describe() for d in result.discrepancies),
        period_start=statement.period_start,
        period_end=statement.period_end,
        credits=result.row_credits,
        debits=result.row_debits,
    )


def build_rows(
    statements: Sequence[Statement | CardStatement],
) -> Iterator[dict[str, Any]]:
    """Every transaction across every statement, in date order, enriched.

    Reuses :func:`query.build_ledger` rather than walking the statements again:
    it already resolves each narration to a merchant against a vocabulary built
    from that statement, and duplicating that here would let the export and the
    `ask` layer drift apart on what a merchant is called.
    """
    ledger = build_ledger(statements)
    certificates = {
        str(statement.source): certify(statement) for statement in statements
    }
    known = {str(statement.source): statement for statement in statements}

    for row in ledger.rows:
        txn = row.transaction
        key = str(row.source)
        statement = known[key]
        certificate = certificates[key]
        debit = txn.direction is Direction.DEBIT
        yield {
            "date": txn.date.isoformat(),
            "value_date": txn.value_date.isoformat() if txn.value_date else "",
            "narration": txn.narration,
            "merchant": row.match.name,
            "category": row.match.category,
            "counterparty_type": row.match.kind.value,
            "debit": rupees(txn.amount) if debit else None,
            "credit": None if debit else rupees(txn.amount),
            "balance": rupees(txn.balance) if txn.balance is not None else None,
            "institution": _institution(statement),
            "account": statement.account_hint,
            "source_file": Path(row.source).name,
            "page": txn.page if txn.page is not None else "",
            "row": txn.row,
            "reconciled": "yes" if certificate.reconciled else "NO",
        }


def _institution(statement: Statement | CardStatement) -> str:
    return statement.issuer if isinstance(statement, CardStatement) else statement.bank


def render_certificates(certificates: Sequence[Certificate]) -> str:
    """The full certificate document, headline verdict first."""
    failed = [c for c in certificates if not c.reconciled]
    stamped = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    if failed:
        headline = (
            f"NOT RECONCILED -- {len(failed)} of {len(certificates)} statement(s) "
            f"did not add up. The rows below may be an incomplete copy of the "
            f"source. Do not file this without reading the failures."
        )
    else:
        headline = (
            f"RECONCILED -- all {len(certificates)} statement(s) add up to the "
            f"paisa against the arithmetic the institution itself published."
        )

    lines = [
        "moneytrail reconciliation certificate",
        f"generated {stamped}",
        "",
        headline,
        "",
    ]
    for certificate in certificates:
        lines.append(certificate.render())
        lines.append("")
    lines.append(
        "Each SHA-256 above is of the exact source bytes that were read. A file "
        "whose digest does not match is not the file this certificate covers."
    )
    return "\n".join(lines)


def export_csv(
    statements: Sequence[Statement | CardStatement], out: Path
) -> list[Certificate]:
    """Write the ledger as CSV and return the certificates.

    ``utf-8-sig``, because Excel on Windows reads a plain UTF-8 CSV as the
    system codepage and turns every rupee sign and every non-ASCII narration
    into mojibake. The BOM is what makes it open correctly by double-click,
    which is how this file will actually be opened.
    """
    out = Path(out)
    rows = list(build_rows(statements))
    with out.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _csv_cell(row[key]) for key in COLUMNS})
    return [certify(statement) for statement in statements]


def _csv_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def export_xlsx(
    statements: Sequence[Statement | CardStatement], out: Path
) -> list[Certificate]:
    """Write a two-sheet workbook: the ledger, and the certificate beside it.

    The certificate lives in the same file as the data deliberately. A proof
    that travels separately from the thing it proves gets detached on the first
    forward, and then the spreadsheet is just another spreadsheet.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    out = Path(out)
    rows = list(build_rows(statements))
    certificates = [certify(statement) for statement in statements]

    book = Workbook()
    sheet = book.active
    sheet.title = "Ledger"
    sheet.append([column.replace("_", " ") for column in COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    money = {"debit", "credit", "balance"}
    for row in rows:
        sheet.append([row[column] for column in COLUMNS])
    for index, column in enumerate(COLUMNS, start=1):
        letter = get_column_letter(index)
        if column in money:
            for cell in sheet[letter][1:]:
                cell.number_format = "#,##0.00"
        sheet.column_dimensions[letter].width = _width(column)

    proof = book.create_sheet("Certificate")
    for line in render_certificates(certificates).splitlines():
        proof.append([line])
    proof["A1"].font = Font(bold=True, size=14)
    proof.column_dimensions["A"].width = 100
    for cell in proof["A"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    book.save(out)
    return certificates


#: Rough character widths. Narration is the only genuinely long field; the rest
#: are short enough that the header sets the floor.
_WIDTHS = {"narration": 52, "merchant": 24, "category": 16, "source_file": 26}


def _width(column: str) -> int:
    return _WIDTHS.get(column, max(12, len(column) + 2))


def write(
    statements: Sequence[Statement | CardStatement], out: Path
) -> list[Certificate]:
    """Export to whichever format ``out``'s suffix names."""
    out = Path(out)
    suffix = out.suffix.lower()
    if suffix == ".xlsx":
        return export_xlsx(statements, out)
    if suffix in {".csv", ".txt", ""}:
        return export_csv(statements, out)
    raise ValueError(
        f"cannot export to {suffix!r} -- use .csv or .xlsx"
    )
